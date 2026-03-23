"""
SIGMA — Pipeline de Notas
Procesa:
  1. Excel de ejes de evaluación (tabla maestra de ponderaciones)
  2. ZIP con archivos XLS de Syscol (un archivo por eje/asignatura por curso)

Salida gold/notas/:
  notas_ejes.csv          — tabla maestra de ejes (cargada del Excel de configuración)
  notas_alumnos.csv       — promedio ponderado por alumno × asignatura × semestre
  notas_cursos.csv        — promedios agregados por curso × asignatura
  notas_detalle_ejes.csv  — promedio por alumno × eje (para análisis de detalle)
  notas_meta.csv          — metadatos del último procesamiento
"""
from __future__ import annotations

import io
import logging
import re
import unicodedata
import zipfile
from pathlib import Path
from datetime import datetime
from typing import Any

import pandas as pd
import numpy as np

log = logging.getLogger("sigma.notas")

NOTA_MINIMA  = 2.0
NOTA_MAXIMA  = 7.0
NOTA_APRUEBA = 4.0


# ══════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════

def _norm(texto: Any) -> str:
    """Normaliza texto: mayúsculas, sin tildes, sin espacios extra."""
    if texto is None:
        return ""
    s = str(texto).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def _parse_nota(v: Any) -> float | None:
    """Convierte un valor a float nota (2.0–7.0) o None si vacío/inválido."""
    if v is None or str(v).strip() in ("", "None", "-", "—"):
        return None
    try:
        f = float(str(v).replace(",", ".").strip())
        if NOTA_MINIMA <= f <= NOTA_MAXIMA:
            return round(f, 1)
        return None
    except (ValueError, TypeError):
        return None


def _parse_nombre_archivo(filename: str) -> dict | None:
    """
    Parsea el nombre de archivo Syscol.
    Formato: {CURSO}-{COD_ASIG}-{N}per_{ID}.xls[x]
    Ej:      1EMA-LENG-1per_234122.xls
             3EMB-MAT-2per_99999.xls
    Retorna dict con keys: curso, cod_asig, semestre_num (1 o 2), o None si no matchea.
    """
    stem = Path(filename).stem  # sin extensión
    # Patrón: CURSO-CODASIG-Nper_ID
    m = re.match(r'^([1-4]EM[A-Z])-([A-Z0-9]+)-([12])per(?:_\d+)?$', stem, re.IGNORECASE)
    if not m:
        return None
    curso      = m.group(1).upper()
    cod_asig   = m.group(2).upper()
    sem_num    = int(m.group(3))
    level      = int(curso[0])
    term_id    = "EXT" if level == 4 else f"S{sem_num}"
    return {
        "curso":      curso,
        "cod_asig":   cod_asig,
        "sem_num":    sem_num,
        "term_id":    term_id,
        "level":      level,
    }


def _extract_pp_from_xls(content: bytes, filename: str) -> pd.DataFrame | None:
    """
    Lee un XLS/XLSX de Syscol y extrae la columna PP (promedio final) por alumno.
    El XLS de Syscol contiene HTML embebido en la primera celda.

    Retorna DataFrame con columnas: nombre, pp
    """
    try:
        # Intentar como HTML embebido en XLS (formato Syscol)
        text = content.decode("latin-1", errors="replace")

        if "<table" in text.lower():
            return _parse_html_table(text, filename)

        # Fallback: intentar con openpyxl si es xlsx real
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            return _parse_openpyxl_sheet(ws, filename)
        except Exception:
            pass

        return None

    except Exception as e:
        log.warning(f"  Error leyendo {filename}: {e}")
        return None


def _parse_html_table(html: str, filename: str) -> pd.DataFrame | None:
    """
    Parsea la tabla HTML embebida en los XLS de Syscol.
    Estructura: L | NOMBRE ALUMNO | 1 | 2 | ... | P | 1 | 2 | ... | P | PP | EX | PA
    Nos interesa: NOMBRE ALUMNO y PP (penúltima columna antes de EX y PA).
    """
    try:
        tables = pd.read_html(io.StringIO(html), header=0)
        if not tables:
            return None

        df = tables[0]
        df.columns = [str(c).strip().upper() for c in df.columns]

        # Identificar columna nombre
        col_nombre = next(
            (c for c in df.columns if "NOMBRE" in c or c == "NOMBRE ALUMNO"),
            None
        )
        if col_nombre is None:
            # Buscar por posición — suele ser la segunda columna
            col_nombre = df.columns[1] if len(df.columns) > 1 else None

        if col_nombre is None:
            log.warning(f"  {filename}: no se encontró columna NOMBRE")
            return None

        # Identificar columna PP
        # En el HTML de Syscol PP aparece como columna con header "PP"
        col_pp = "PP" if "PP" in df.columns else None

        if col_pp is None:
            # Buscar penúltima columna numérica antes de PA y EX
            # Syscol: ... | PP | EX | PA  (últimas 3)
            numeric_cols = [c for c in df.columns
                            if c not in (col_nombre, "L", "EX", "PA", "PR", "CT", "%", "MIN", "MAX", "MD", "DE")]
            if len(numeric_cols) >= 1:
                col_pp = numeric_cols[-1]

        if col_pp is None:
            log.warning(f"  {filename}: no se encontró columna PP")
            return None

        # Filtrar solo filas de alumnos (excluir PR=Promedios, CT=Cantidad, etc.)
        col_lista = df.columns[0]  # columna L con número de lista
        df_alum = df[
            df[col_lista].astype(str).str.match(r'^\d{1,2}$', na=False)
        ].copy()

        if df_alum.empty:
            log.warning(f"  {filename}: sin filas de alumnos")
            return None

        result = pd.DataFrame({
            "nombre": df_alum[col_nombre].astype(str).str.strip().str.upper(),
            "pp":     df_alum[col_pp].apply(_parse_nota),
        })

        result = result[result["nombre"].str.len() > 3].copy()
        log.info(f"  {filename}: {len(result)} alumnos, col_pp='{col_pp}'")
        return result

    except Exception as e:
        log.warning(f"  {filename}: error parseando HTML: {e}")
        return None


def _parse_openpyxl_sheet(ws, filename: str) -> pd.DataFrame | None:
    """Fallback para XLSX reales (no HTML embebido)."""
    try:
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
        if not rows:
            return None

        # Header en fila 1
        header = [str(v).strip().upper() if v else f"COL{i}"
                  for i, v in enumerate(rows[0])]
        data   = rows[1:]

        df = pd.DataFrame(data, columns=header)

        col_nombre = next((c for c in header if "NOMBRE" in c), header[1] if len(header) > 1 else None)
        col_pp     = "PP" if "PP" in header else None

        if col_nombre is None or col_pp is None:
            return None

        result = pd.DataFrame({
            "nombre": df[col_nombre].astype(str).str.strip().str.upper(),
            "pp":     df[col_pp].apply(_parse_nota),
        })
        return result[result["nombre"].str.len() > 3].copy()

    except Exception as e:
        log.warning(f"  {filename}: error openpyxl: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════
# CARGA DE EJES (tabla maestra)
# ══════════════════════════════════════════════════════════════════════

def load_ejes(excel_path: str | Path) -> pd.DataFrame:
    """
    Carga el Excel de ejes de evaluación y lo valida.
    Retorna DataFrame normalizado listo para usar como tabla maestra.
    """
    import openpyxl
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de ejes: {path}")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    wb.close()  # Cerrar explícitamente para liberar el archivo en Windows

    df = pd.DataFrame(rows[1:], columns=rows[0])

    # Normalizar columnas requeridas
    required = ["academic_year", "level", "specialty", "term_id",
                "cod_asignatura", "asignatura", "cod_eje", "eje_evaluacion", "ponderacion_eje"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas en el Excel de ejes: {missing}")

    df["level"]           = pd.to_numeric(df["level"], errors="coerce").astype("Int64")
    df["ponderacion_eje"] = pd.to_numeric(df["ponderacion_eje"], errors="coerce")
    df["specialty"]       = df["specialty"].astype(str).str.strip().str.upper()
    df["term_id"]         = df["term_id"].astype(str).str.strip().str.upper()
    df["cod_asignatura"]  = df["cod_asignatura"].astype(str).str.strip().str.upper()
    df["cod_eje"]         = df["cod_eje"].astype(str).str.strip().str.upper()
    df["asignatura"]      = df["asignatura"].astype(str).str.strip().str.upper()
    df["eje_evaluacion"]  = df["eje_evaluacion"].astype(str).str.strip().str.upper()

    # Validar ponderaciones
    problemas = []
    for (lvl, spec, asig, term), grp in df.groupby(["level", "specialty", "asignatura", "term_id"]):
        total = grp["ponderacion_eje"].sum()
        if abs(total - 100) > 1:
            problemas.append(f"{lvl}° {spec} | {asig} | {term}: suma={total:.1f}%")

    if problemas:
        log.warning(f"  Ponderaciones != 100% en: {problemas}")

    log.info(f"✅ Ejes cargados: {len(df)} filas, {df['asignatura'].nunique()} asignaturas")
    return df.reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════
# PROCESAMIENTO DEL ZIP DE NOTAS
# ══════════════════════════════════════════════════════════════════════

def process_zip_notas(
    zip_path: str | Path,
    df_ejes: pd.DataFrame,
    gold_dir: str | Path = "data/gold/notas",
) -> dict:
    """
    Procesa un ZIP con archivos XLS de Syscol.
    Para cada archivo:
      1. Parsea el nombre → curso, cod_asig, term_id
      2. Busca el eje correspondiente en df_ejes
      3. Extrae la columna PP por alumno
      4. Acumula en gold/notas/_raw_ejes_acumulado.csv (deduplicado)
    Luego recalcula los promedios ponderados por asignatura.

    Retorna dict con DataFrames resultado.
    """
    gold_dir = Path(gold_dir)
    gold_dir.mkdir(parents=True, exist_ok=True)

    zip_path = Path(zip_path)
    if not zip_path.exists():
        raise FileNotFoundError(f"ZIP no encontrado: {zip_path}")

    log.info(f"Procesando ZIP: {zip_path.name}")

    # ── Leer archivos del ZIP ──────────────────────────────────────────
    nuevos_registros = []
    archivos_ok = 0
    archivos_skip = 0

    with zipfile.ZipFile(zip_path, "r") as zf:
        for member in zf.namelist():
            fname = Path(member).name
            if not fname.lower().endswith((".xls", ".xlsx")):
                continue
            if fname.startswith("~") or fname.startswith("."):
                continue

            parsed = _parse_nombre_archivo(fname)
            if parsed is None:
                log.warning(f"  SKIP (nombre no reconocido): {fname}")
                archivos_skip += 1
                continue

            curso    = parsed["curso"]
            cod_asig = parsed["cod_asig"]
            term_id  = parsed["term_id"]
            level    = parsed["level"]

            # Verificar que el eje existe en tabla maestra
            mask_eje = (
                (df_ejes["cod_asignatura"] == cod_asig) &
                (df_ejes["term_id"] == term_id)
            )
            if not mask_eje.any():
                # Intentar buscar por nombre aproximado en caso de que el código
                # no esté en el Excel (archivo extra que no esperábamos)
                log.warning(f"  SKIP (eje no en tabla maestra): {fname} → {cod_asig}/{term_id}")
                archivos_skip += 1
                continue

            # Leer contenido
            with zf.open(member) as f:
                content = f.read()

            df_pp = _extract_pp_from_xls(content, fname)
            if df_pp is None or df_pp.empty:
                log.warning(f"  SKIP (sin datos): {fname}")
                archivos_skip += 1
                continue

            # Agregar metadatos
            df_pp["curso"]    = curso
            df_pp["cod_asig"] = cod_asig
            df_pp["term_id"]  = term_id
            df_pp["level"]    = level
            df_pp["fuente"]   = fname

            nuevos_registros.append(df_pp)
            archivos_ok += 1

    log.info(f"  Archivos OK: {archivos_ok} | Saltados: {archivos_skip}")

    if not nuevos_registros:
        log.warning("  Sin datos procesados del ZIP")
        return {"alumnos": pd.DataFrame(), "cursos": pd.DataFrame(), "detalle": pd.DataFrame()}

    df_nuevos = pd.concat(nuevos_registros, ignore_index=True)

    # ── Acumular con gold existente + deduplicar ───────────────────────
    raw_path = gold_dir / "_raw_ejes_acumulado.csv"
    if raw_path.exists():
        try:
            df_existing = pd.read_csv(raw_path, encoding="utf-8")
            df_raw = pd.concat([df_existing, df_nuevos], ignore_index=True)
            antes = len(df_raw)
            # Clave única: nombre + curso + cod_asig + term_id
            df_raw = df_raw.drop_duplicates(
                subset=["nombre", "curso", "cod_asig", "term_id"],
                keep="last"
            )
            log.info(f"  Acumulado: {len(df_raw)} registros ({antes - len(df_raw)} duplicados eliminados)")
        except Exception as e:
            log.warning(f"  No se pudo cargar raw existente: {e}")
            df_raw = df_nuevos
    else:
        df_raw = df_nuevos

    df_raw.to_csv(raw_path, index=False, encoding="utf-8")

    # ── Calcular promedios ponderados ──────────────────────────────────
    resultados = _calcular_promedios(df_raw, df_ejes)

    # ── Guardar gold ───────────────────────────────────────────────────
    resultados["detalle"].to_csv(gold_dir / "notas_detalle_ejes.csv",  index=False, encoding="utf-8")
    resultados["alumnos"].to_csv(gold_dir / "notas_alumnos.csv",       index=False, encoding="utf-8")
    resultados["cursos"].to_csv( gold_dir / "notas_cursos.csv",        index=False, encoding="utf-8")

    # Guardar tabla maestra de ejes
    df_ejes.to_csv(gold_dir / "notas_ejes.csv", index=False, encoding="utf-8")

    # Metadatos
    meta = pd.DataFrame([{
        "generado":       datetime.now().isoformat(),
        "archivos_ok":    archivos_ok,
        "archivos_skip":  archivos_skip,
        "alumnos":        len(resultados["alumnos"]),
        "asignaturas":    resultados["alumnos"]["cod_asig"].nunique() if not resultados["alumnos"].empty else 0,
        "cursos":         resultados["alumnos"]["curso"].nunique() if not resultados["alumnos"].empty else 0,
    }])
    meta.to_csv(gold_dir / "notas_meta.csv", index=False, encoding="utf-8")

    log.info(f"✅ Notas gold guardado en {gold_dir}")
    log.info(f"   Alumnos: {len(resultados['alumnos'])} | Cursos: {resultados['alumnos']['curso'].nunique() if not resultados['alumnos'].empty else 0}")

    return resultados


def _calcular_promedios(df_raw: pd.DataFrame, df_ejes: pd.DataFrame) -> dict:
    """
    Calcula promedios ponderados por asignatura para cada alumno.

    Lógica:
      - Para cada alumno × curso × asignatura × term_id:
        nota_asig = Σ (pp_eje × ponderacion_eje/100)
        Solo si hay al menos 1 eje con nota disponible.
        Si falta algún eje, la ponderación se redistribuye entre los disponibles.
    """
    if df_raw.empty:
        return {"alumnos": pd.DataFrame(), "cursos": pd.DataFrame(), "detalle": pd.DataFrame()}

    # Unir con tabla maestra para obtener ponderaciones
    df_merged = df_raw.merge(
        df_ejes[["cod_asignatura", "cod_eje", "term_id", "asignatura",
                 "ponderacion_eje", "specialty", "level"]].rename(
            columns={"cod_asignatura": "cod_asig"}),
        on=["cod_asig", "term_id"],
        how="left"
    )

    # ── Detalle por eje ────────────────────────────────────────────────
    detalle = df_merged[[
        "nombre", "curso", "cod_asig", "asignatura",
        "cod_eje", "eje_evaluacion" if "eje_evaluacion" in df_merged.columns else "cod_eje",
        "term_id", "pp", "ponderacion_eje"
    ]].copy()

    detalle = detalle.rename(columns={
        "pp": "nota_eje",
        "eje_evaluacion": "eje_nombre" if "eje_evaluacion" in df_merged.columns else "cod_eje"
    })

    # ── Promedio ponderado por alumno × asignatura ─────────────────────
    registros_alumnos = []

    grupos = df_merged.groupby(["nombre", "curso", "cod_asig", "term_id"])
    for (nombre, curso, cod_asig, term_id), grp in grupos:
        # Filtrar ejes con nota disponible
        con_nota = grp[grp["pp"].notna()].copy()

        if con_nota.empty:
            continue

        # Redistribuir ponderación entre ejes con nota
        pond_total_disponible = con_nota["ponderacion_eje"].sum()
        if pond_total_disponible <= 0:
            continue

        # Calcular promedio ponderado redistribuido
        nota_ponderada = (
            (con_nota["pp"] * con_nota["ponderacion_eje"]).sum()
            / pond_total_disponible
        )
        nota_ponderada = round(nota_ponderada, 1)

        asignatura = grp["asignatura"].iloc[0] if "asignatura" in grp.columns else cod_asig
        level      = grp["level"].iloc[0] if "level" in grp.columns else None
        specialty  = grp["specialty"].iloc[0] if "specialty" in grp.columns else None

        n_ejes_total     = len(grp["cod_eje"].unique()) if "cod_eje" in grp.columns else len(grp)
        n_ejes_con_nota  = len(con_nota["cod_eje"].unique()) if "cod_eje" in con_nota.columns else len(con_nota)

        registros_alumnos.append({
            "nombre":           nombre,
            "curso":            curso,
            "level":            level,
            "specialty":        specialty,
            "cod_asig":         cod_asig,
            "asignatura":       asignatura,
            "term_id":          term_id,
            "nota":             nota_ponderada,
            "aprobado":         1 if nota_ponderada >= NOTA_APRUEBA else 0,
            "n_ejes_total":     n_ejes_total,
            "n_ejes_con_nota":  n_ejes_con_nota,
            "completo":         1 if n_ejes_con_nota == n_ejes_total else 0,
            "alerta":           _alerta_nota(nota_ponderada, n_ejes_con_nota, n_ejes_total),
        })

    df_alumnos = pd.DataFrame(registros_alumnos)

    if df_alumnos.empty:
        return {"alumnos": df_alumnos, "cursos": pd.DataFrame(), "detalle": detalle}

    # ── Promedio por curso × asignatura ───────────────────────────────
    df_cursos = (
        df_alumnos.groupby(["curso", "cod_asig", "asignatura", "term_id"], as_index=False)
        .agg(
            prom_curso    = ("nota",     "mean"),
            n_alumnos     = ("nombre",   "count"),
            n_aprobados   = ("aprobado", "sum"),
            nota_max      = ("nota",     "max"),
            nota_min      = ("nota",     "min"),
        )
    )
    df_cursos["prom_curso"]   = df_cursos["prom_curso"].round(1)
    df_cursos["pct_aprobado"] = (df_cursos["n_aprobados"] / df_cursos["n_alumnos"] * 100).round(1)
    df_cursos["alerta_curso"] = df_cursos["prom_curso"].apply(
        lambda v: "CRITICO" if v < 4.5 else ("ATENCION" if v < 5.0 else "OK")
    )

    return {
        "alumnos": df_alumnos,
        "cursos":  df_cursos,
        "detalle": detalle,
    }


def _alerta_nota(nota: float, ejes_con_nota: int, ejes_total: int) -> str:
    if nota < NOTA_APRUEBA:
        return "REPROBADO"
    if nota < 4.5:
        return "EN_RIESGO"
    if ejes_con_nota < ejes_total:
        return "INCOMPLETO"
    return "OK"


# ══════════════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL run()
# ══════════════════════════════════════════════════════════════════════

def run(
    zip_path: str | Path | None = None,
    ejes_path: str | Path | None = None,
    gold_dir: str | Path = "data/gold/notas",
) -> dict:
    """
    Punto de entrada principal del pipeline.

    Puede llamarse con:
      - Solo ejes_path → actualiza tabla maestra
      - Solo zip_path  → procesa notas (carga ejes desde gold existente)
      - Ambos          → actualiza ejes Y procesa notas
    """
    gold_dir = Path(gold_dir)
    gold_dir.mkdir(parents=True, exist_ok=True)

    resultado = {}

    # Cargar o actualizar tabla maestra de ejes
    if ejes_path is not None:
        log.info("Cargando tabla maestra de ejes...")
        df_ejes = load_ejes(ejes_path)
        df_ejes.to_csv(gold_dir / "notas_ejes.csv", index=False, encoding="utf-8")
        resultado["ejes"] = df_ejes
        log.info(f"✅ Ejes guardados: {len(df_ejes)} filas")
    else:
        # Intentar cargar desde gold existente
        ejes_gold = gold_dir / "notas_ejes.csv"
        if ejes_gold.exists():
            df_ejes = pd.read_csv(ejes_gold, encoding="utf-8")
            resultado["ejes"] = df_ejes
        else:
            log.error("No hay tabla de ejes disponible. Sube el Excel de ejes primero.")
            return resultado

    # Procesar ZIP de notas
    if zip_path is not None:
        log.info("Procesando ZIP de notas...")
        resultado.update(process_zip_notas(zip_path, resultado["ejes"], gold_dir))

    return resultado


# ══════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    parser = argparse.ArgumentParser(description="SIGMA — Pipeline de Notas")
    parser.add_argument("--ejes",  help="Excel de ejes de evaluación")
    parser.add_argument("--zip",   help="ZIP con archivos XLS de Syscol")
    parser.add_argument("--gold",  default="data/gold/notas", help="Directorio gold")
    args = parser.parse_args()

    r = run(
        zip_path  = args.zip,
        ejes_path = args.ejes,
        gold_dir  = args.gold,
    )
    print(f"Resultado: {list(r.keys())}")