"""
SIGMA — Módulo de Reportes
Genera reportes en PDF y Excel para cada módulo.

Uso desde app.py:
    from sigma_reports import generate_pdf_ejecutivo, generate_excel_ejecutivo, ...
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd

# ── ReportLab ────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.platypus import Image as RLImage

# Importar gráficos PDF (matplotlib)
try:
    from sigma_pdf_charts import (
        pdf_chart_genero_nivel, pdf_chart_especialidad_resumen,
        pdf_chart_heatmap, pdf_chart_cursos,
        pdf_chart_des_nivel, pdf_chart_des_especialidad, pdf_chart_des_perfil,
    )
    _CHARTS_OK = True
    print("[SIGMA] sigma_pdf_charts cargado OK")
except Exception as e:
    _CHARTS_OK = False
    print(f"[SIGMA] sigma_pdf_charts NO cargado: {type(e).__name__}: {e}")
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ── openpyxl ─────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────
# PALETA SIGMA — EJECUTIVA (fondo blanco, acentos sobrios)
# ─────────────────────────────────────────────────────────────────────
C_BG       = colors.white
C_PANEL    = colors.white
C_WHITE    = colors.white

C_CYAN     = colors.HexColor("#2563eb")   # azul corporativo
C_GREEN    = colors.HexColor("#16a34a")   # verde sobrio
C_AMBER    = colors.HexColor("#d97706")   # ámbar sobrio
C_RED      = colors.HexColor("#dc2626")   # rojo sobrio
C_NAVY     = colors.HexColor("#1e3a5f")   # azul oscuro header

C_TEXT     = colors.HexColor("#111827")   # texto principal
C_MUTED    = colors.HexColor("#6b7280")   # texto secundario
C_BORDER   = colors.HexColor("#e5e7eb")   # bordes tabla
C_ROW_ALT  = colors.HexColor("#f9fafb")   # fila alterna

SPEC_COLORS = {
    "TELECOM":     colors.HexColor("#2563eb"),
    "ELECTRONICA": colors.HexColor("#16a34a"),
    "MECANICA":    colors.HexColor("#d97706"),
    "COMUN":       colors.HexColor("#7c3aed"),
}

# ─────────────────────────────────────────────────────────────────────
# ESTILOS PDF
# ─────────────────────────────────────────────────────────────────────
def _build_styles():
    titulo = ParagraphStyle(
        "SigmaTitulo", fontName="Helvetica-Bold", fontSize=16,
        textColor=C_TEXT, spaceAfter=2, leading=20,
    )
    subtitulo = ParagraphStyle(
        "SigmaSubtitulo", fontName="Helvetica", fontSize=9,
        textColor=C_MUTED, spaceAfter=8,
    )
    seccion = ParagraphStyle(
        "SigmaSeccion", fontName="Helvetica-Bold", fontSize=9,
        textColor=C_MUTED, spaceBefore=12, spaceAfter=4,
    )
    normal = ParagraphStyle(
        "SigmaNormal", fontName="Helvetica", fontSize=9,
        textColor=C_TEXT, spaceAfter=4, leading=13,
    )
    kpi_val = ParagraphStyle(
        "SigmaKpiVal", fontName="Helvetica-Bold", fontSize=22,
        textColor=C_TEXT, alignment=1, leading=26,
    )
    kpi_lbl = ParagraphStyle(
        "SigmaKpiLbl", fontName="Helvetica", fontSize=7,
        textColor=C_MUTED, alignment=1, leading=9,
    )
    tabla_header = ParagraphStyle(
        "SigmaTablaH", fontName="Helvetica-Bold", fontSize=8,
        textColor=C_TEXT,
    )
    tabla_cell = ParagraphStyle(
        "SigmaTablaC", fontName="Helvetica", fontSize=8,
        textColor=C_TEXT,
    )
    footer = ParagraphStyle(
        "SigmaFooter", fontName="Helvetica", fontSize=7,
        textColor=C_MUTED, alignment=1,
    )
    return dict(
        titulo=titulo, subtitulo=subtitulo, seccion=seccion,
        normal=normal, kpi_val=kpi_val, kpi_lbl=kpi_lbl,
        tabla_header=tabla_header, tabla_cell=tabla_cell, footer=footer,
    )


def _hr(color=C_CYAN, thickness=0.5):
    return HRFlowable(width="100%", thickness=thickness, color=color, spaceAfter=8, spaceBefore=4)


def _kpi_table(kpis: list[tuple]) -> Table:
    """kpis = [(label, value, color)] o [(label, value, color, sub)]
    sub: texto pequeño opcional que aparece entre el valor y el label (ej: porcentaje).
    Estilo ejecutivo: fondo blanco, número grande en color, borde inferior de color."""
    n     = len(kpis)
    col_w = (A4[0] - 3 * cm) / n

    val_row = []
    sub_row = []
    lbl_row = []
    has_sub = any(len(k) > 3 and k[3] for k in kpis)

    style_cmds = [
        ("BACKGROUND",    (0, 0), (-1, -1), colors.white),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEAFTER",     (0, 0), (-2, -1), 0.5, C_BORDER),
        ("BOX",           (0, 0), (-1, -1), 0.5, C_BORDER),
    ]

    for i, kpi in enumerate(kpis):
        label, value, color = kpi[0], kpi[1], kpi[2]
        sub = kpi[3] if len(kpi) > 3 else ""

        val_row.append(Paragraph(str(value), ParagraphStyle(
            f"kv{i}", fontName="Helvetica-Bold", fontSize=22,
            textColor=color, alignment=1, leading=26,
        )))
        sub_row.append(Paragraph(str(sub) if sub else "", ParagraphStyle(
            f"ks{i}", fontName="Helvetica", fontSize=7,
            textColor=color, alignment=1, leading=9,
        )))
        lbl_row.append(Paragraph(str(label).upper(), ParagraphStyle(
            f"kl{i}", fontName="Helvetica", fontSize=6.5,
            textColor=C_MUTED, alignment=1, leading=9,
        )))
        style_cmds.append(("LINEBELOW", (i, 0), (i, 0), 2.5, color))

    if has_sub:
        data      = [val_row, sub_row, lbl_row]
        row_h     = [30, 12, 14]
    else:
        data      = [val_row, lbl_row]
        row_h     = [32, 16]

    t = Table(data, colWidths=[col_w] * n, rowHeights=row_h)
    t.setStyle(TableStyle(style_cmds))
    return t


def _sigma_header(story: list, titulo: str, subtitulo: str, corte: str):
    """Header ejecutivo: fondo blanco, logo azul, línea divisoria cian."""
    PAGE_W = A4[0] - 3 * cm

    logo_style = ParagraphStyle(
        "SigmaLogo", fontName="Helvetica-Bold", fontSize=20,
        textColor=C_NAVY, leading=24, alignment=0,
    )
    tag_style = ParagraphStyle(
        "SigmaTag", fontName="Helvetica", fontSize=7.5,
        textColor=C_MUTED, leading=10, alignment=0,
    )
    tit_style = ParagraphStyle(
        "SigmaHdrTit", fontName="Helvetica-Bold", fontSize=13,
        textColor=C_TEXT, leading=17, alignment=2,
    )
    meta_style = ParagraphStyle(
        "SigmaHdrMeta", fontName="Helvetica", fontSize=8,
        textColor=C_MUTED, leading=11, alignment=2,
    )

    left_col  = [[Paragraph("SIGMA", logo_style)],
                 [Paragraph("Sistema Integrado de Gestión y Monitoreo Académico", tag_style)]]
    right_col = [[Paragraph(titulo, tit_style)],
                 [Paragraph(f"Corte: {corte}  ·  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", meta_style)]]

    left_t = Table(left_col, colWidths=[PAGE_W * 0.5], rowHeights=[26, 14])
    left_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.white),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("VALIGN",        (0,0),(-1,-1), "BOTTOM"),
    ]))
    right_t = Table(right_col, colWidths=[PAGE_W * 0.5], rowHeights=[26, 14])
    right_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.white),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("VALIGN",        (0,0),(-1,-1), "BOTTOM"),
    ]))

    header_t = Table([[left_t, right_t]],
                     colWidths=[PAGE_W * 0.5, PAGE_W * 0.5], rowHeights=[48])
    header_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), colors.white),
        ("TOPPADDING",    (0,0),(-1,-1), 8),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (0,0),(-1,-1), 0),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("LINEBELOW",     (0,0),(-1,-1), 1.5, C_CYAN),
    ]))

    story.append(header_t)
    story.append(Spacer(1, 0.3 * cm))


def _df_to_table(df: pd.DataFrame, col_widths: Optional[list] = None,
                 row_colors: bool = True) -> Table:
    """Tabla ejecutiva: fondo blanco, header azul oscuro, zebra gris suave."""
    s = _build_styles()
    hdr_style = ParagraphStyle(
        "th", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        "td", fontName="Helvetica", fontSize=8, textColor=C_TEXT,
    )
    headers = [Paragraph(str(c), hdr_style) for c in df.columns]
    rows = [headers]
    for _, row in df.iterrows():
        rows.append([Paragraph(str(v) if pd.notna(v) else "", cell_style)
                     for v in row])

    if col_widths is None:
        avail = A4[0] - 3 * cm
        col_widths = [avail / len(df.columns)] * len(df.columns)

    t = Table(rows, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        # Header: azul oscuro, texto blanco
        ("BACKGROUND",    (0, 0), (-1, 0),  C_NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        # Bordes: solo horizontales finos, gris claro
        ("LINEBELOW",     (0, 0), (-1, -1), 0.4, C_BORDER),
        ("LINEABOVE",     (0, 0), (-1, 0),  0.4, C_BORDER),
        ("BOX",           (0, 0), (-1, -1), 0.4, C_BORDER),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ]
    if row_colors:
        for i in range(1, len(rows)):
            if i % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), C_ROW_ALT))
            else:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.white))
    t.setStyle(TableStyle(style_cmds))
    return t


def _sigma_footer(story: list):
    story.append(Spacer(1, 0.4 * cm))
    footer_t = Table(
        [[Paragraph(
            f"SIGMA  ·  Matrícula  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Documento de uso interno",
            ParagraphStyle("ft", fontName="Helvetica", fontSize=7, textColor=C_MUTED, alignment=1),
        )]],
        colWidths=[A4[0] - 3 * cm],
        rowHeights=[18],
    )
    footer_t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), colors.white),
        ("LINEABOVE",     (0, 0), (-1, -1), 0.5, C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(footer_t)


def _chart_block(story: list, img_bytes: bytes, texto: str,
                  width_cm: float = 17, caption: str = ""):
    """Inserta imagen matplotlib + párrafo de texto analítico en el story."""
    if not img_bytes:
        return
    from io import BytesIO
    img = RLImage(BytesIO(img_bytes), width=width_cm*cm,
                  height=width_cm*cm*0.38)
    story.append(img)
    if texto:
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(texto, ParagraphStyle(
            "insight", fontName="Helvetica", fontSize=8,
            textColor=colors.HexColor("#374151"),
            backColor=colors.HexColor("#f9fafb"),
            borderPad=6, leading=12, spaceAfter=4,
            leftIndent=6, rightIndent=6,
            borderWidth=0, borderColor=colors.HexColor("#e5e7eb"),
        )))
    story.append(Spacer(1, 0.35*cm))


def _section_title(text: str, color=C_CYAN) -> Table:
    """Título de sección ejecutivo: texto gris oscuro en caps, línea azul a la izquierda."""
    t = Table(
        [[Paragraph(text.upper(), ParagraphStyle(
            "st", fontName="Helvetica-Bold", fontSize=8,
            textColor=C_MUTED, leading=11, spaceAfter=0,
        ))]],
        colWidths=[A4[0] - 3 * cm],
        rowHeights=[20],
    )
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), colors.white),
        ("LINEBEFORE",    (0, 0), (0, -1),  2.5, color),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.4, C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    return t


# ─────────────────────────────────────────────────────────────────────
# HELPERS ROBUSTOS DE COLUMNAS
# ─────────────────────────────────────────────────────────────────────
def _norm_sexo(series: pd.Series) -> pd.Series:
    """Normaliza sexo a 'M' o 'F' independiente del formato fuente."""
    s = series.astype(str).str.upper().str.strip()
    return s.map(lambda v: "M" if v in ("M", "MASCULINO") else ("F" if v in ("F", "FEMENINO") else ""))

def _count_hombres(df: pd.DataFrame) -> int:
    if "sexo" not in df.columns:
        return 0
    return int((_norm_sexo(df["sexo"]) == "M").sum())

def _count_mujeres(df: pd.DataFrame) -> int:
    if "sexo" not in df.columns:
        return 0
    return int((_norm_sexo(df["sexo"]) == "F").sum())

def _count_extranjeros(df: pd.DataFrame) -> int:
    if "is_extranjero" in df.columns:
        return int(df["is_extranjero"].sum())
    if "nacionalidad" in df.columns:
        return int((df["nacionalidad"].astype(str).str.upper().str.strip() != "CHILENA").sum())
    return 0


# ═════════════════════════════════════════════════════════════════════
# 1. REPORTE EJECUTIVO GENERAL — PDF
# ═════════════════════════════════════════════════════════════════════
def generate_pdf_ejecutivo(
    metrics: dict,
    df_current: pd.DataFrame,
    df_master_metrics: Optional[pd.DataFrame],
    df_diff: Optional[pd.DataFrame],
    corte: str,
    df_desiste: Optional[pd.DataFrame] = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    s = _build_styles()
    story = []

    # Header
    _sigma_header(story, "Resumen Ejecutivo de Matrícula", "", corte)

    # KPIs principales
    story.append(_section_title("Indicadores principales"))
    matriculados = metrics.get("matriculados", 0)
    desiste      = metrics.get("desiste", 0)
    total_esp    = metrics.get("total_esperado", 0)
    h            = metrics.get("sexo_h", 0)
    m            = metrics.get("sexo_m", 0)
    repitentes   = metrics.get("repitentes", 0)
    extranjeros  = metrics.get("extranjeros_pct", 0)

    story.append(_kpi_table([
        ("Matriculados",    matriculados, C_CYAN),
        ("Desiste",         desiste,      C_RED),
        ("Total esperado",  total_esp,    C_AMBER),
        ("Hombres",         h,            C_CYAN),
        ("Mujeres",         m,            C_GREEN),
        ("Repitentes",      repitentes,   C_AMBER),
    ]))
    story.append(Spacer(1, 0.4 * cm))

    # Distribución por especialidad
    if df_current is not None and not df_current.empty:
        story.append(_section_title("Distribución por especialidad"))
        _df_sp_work = df_current.copy()
        _df_sp_work["_sx"] = _norm_sexo(_df_sp_work["sexo"]) if "sexo" in _df_sp_work.columns else ""
        df_sp = (
            _df_sp_work.groupby("specialty")
            .agg(
                Total   = ("rut_norm", "count"),
                Hombres = ("_sx", lambda x: (x == "M").sum()),
                Mujeres = ("_sx", lambda x: (x == "F").sum()),
            )
            .reset_index()
            .sort_values("Total", ascending=False)
            .rename(columns={"specialty": "Especialidad"})
        )
        df_sp["% del total"] = (df_sp["Total"] / len(df_current) * 100).round(1).astype(str) + "%"
        story.append(_df_to_table(df_sp, col_widths=[5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]))
        story.append(Spacer(1, 0.4 * cm))

    # Cambios vs corte anterior
    if df_diff is not None and not df_diff.empty:
        story.append(_section_title("Cambios vs corte anterior"))
        resumen = df_diff["change_type"].value_counts().reset_index()
        resumen.columns = ["Tipo de cambio", "Cantidad"]
        story.append(_df_to_table(resumen, col_widths=[9*cm, 6*cm]))
        story.append(Spacer(1, 0.3 * cm))

        # Tabla de removidos
        removidos = df_diff[df_diff["change_type"] == "REMOVED"].copy()
        if not removidos.empty:
            story.append(_section_title(f"Alumnos removidos ({len(removidos)})"))
            cols = [c for c in ["rut_norm", "nombre", "course_code", "specialty"] if c in removidos.columns]
            story.append(_df_to_table(
                removidos[cols].reset_index(drop=True),
                col_widths=[3.5*cm, 6.5*cm, 2.5*cm, 2.5*cm],
            ))

    # ── Gráficos ──────────────────────────────────────────────────────
    if _CHARTS_OK and df_current is not None and not df_current.empty:
        story.append(PageBreak())
        _sigma_header(story, "Análisis Visual", "", corte)

        story.append(_section_title("Matrícula por nivel y género"))
        img, txt = pdf_chart_genero_nivel(df_current)
        _chart_block(story, img, txt)

        story.append(_section_title("Matrícula por especialidad"))
        img, txt = pdf_chart_especialidad_resumen(df_current, df_desiste)
        _chart_block(story, img, txt)

    _sigma_footer(story)
    doc.build(story)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════
# 2. REPORTE DESISTIMIENTO — PDF
# ═════════════════════════════════════════════════════════════════════
def generate_pdf_desistimiento(
    df_desiste: pd.DataFrame,
    df_desiste_prev: Optional[pd.DataFrame],
    corte: str,
    df_current: Optional[pd.DataFrame] = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    s = _build_styles()
    story = []

    _sigma_header(story, "Análisis de Desistimiento", "", corte)

    dd = df_desiste.copy()
    dd["specialty"]    = dd.get("specialty", pd.Series(dtype=str)).fillna("SIN DATOS").str.upper()
    dd["sexo"]         = dd.get("sexo",      pd.Series(dtype=str)).fillna("").str.upper()
    dd["is_extranjero"]= ~dd.get("nacionalidad", pd.Series(dtype=str)).fillna("").str.upper().isin(
                            ["CHILENA", "CHILENO", "CHILE", ""])
    dd["is_repeat"]    = dd.get("is_repeat", pd.Series(dtype=bool)).fillna(False).astype(bool)
    dd["level"]        = pd.to_numeric(dd.get("level", pd.Series(dtype=int)), errors="coerce").fillna(0).astype(int)

    total   = len(dd)
    n_h     = int((dd["sexo"] == "MASCULINO").sum())
    n_m     = int((dd["sexo"] == "FEMENINO").sum())
    n_ext   = int(dd["is_extranjero"].sum())
    n_rep   = int(dd["is_repeat"].sum())
    n_norep = total - n_rep

    nuevos = pd.DataFrame()
    if df_desiste_prev is not None:
        nuevos = dd[~dd["rut_norm"].isin(df_desiste_prev["rut_norm"])].copy()

    # KPIs
    story.append(_section_title("Indicadores de desistimiento"))
    story.append(_kpi_table([
        ("Total desiste",   total,   C_RED),
        ("Nuevos corte",    len(nuevos), C_AMBER),
        ("Hombres",         n_h,     C_CYAN),
        ("Mujeres",         n_m,     C_GREEN),
        ("Repitentes",      n_rep,   C_RED),
        ("No repitentes",   n_norep, C_GREEN),
        ("Extranjeros",     n_ext,   C_AMBER),
    ]))
    story.append(Spacer(1, 0.4*cm))

    # Análisis repitentes
    pct_rep = round(n_rep / total * 100, 1) if total else 0
    story.append(_section_title("Análisis de repitencia"))
    story.append(Paragraph(
        f"De los {total} desistes, <b>{n_rep} son repitentes ({pct_rep}%)</b>. "
        f"Considerando que los repitentes representan solo el 4.6% de la matrícula total, "
        f"su tasa de desistimiento es significativamente mayor al promedio.",
        s["normal"],
    ))
    story.append(Spacer(1, 0.3*cm))

    # Por especialidad
    story.append(_section_title("Desistimiento por especialidad"))
    df_sp = (
        dd.groupby("specialty")
        .agg(
            Total      = ("rut_norm", "count"),
            Hombres    = ("sexo",      lambda x: (x == "MASCULINO").sum()),
            Mujeres    = ("sexo",      lambda x: (x == "FEMENINO").sum()),
            Repitentes = ("is_repeat", "sum"),
        )
        .reset_index()
        .sort_values("Total", ascending=False)
        .rename(columns={"specialty": "Especialidad"})
    )
    df_sp["% del total"] = (df_sp["Total"] / total * 100).round(1).astype(str) + "%"
    story.append(_df_to_table(
        df_sp, col_widths=[4.5*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm],
    ))
    story.append(Spacer(1, 0.3*cm))

    # Por nivel
    story.append(_section_title("Desistimiento por nivel"))
    df_niv = (
        dd.groupby("level")
        .agg(
            Total      = ("rut_norm", "count"),
            Repitentes = ("is_repeat", "sum"),
        )
        .reset_index()
        .sort_values("level")
        .rename(columns={"level": "Nivel"})
    )
    df_niv["Nivel"] = df_niv["Nivel"].astype(str) + "° Medio"
    df_niv["% del total"] = (df_niv["Total"] / total * 100).round(1).astype(str) + "%"
    story.append(_df_to_table(df_niv, col_widths=[4*cm, 3*cm, 3*cm, 3*cm]))
    story.append(Spacer(1, 0.3*cm))

    # Top comunas
    story.append(_section_title("Desistimiento por comuna"))
    df_com = (
        dd.groupby("comuna")
        .agg(Total=("rut_norm", "count"))
        .reset_index()
        .sort_values("Total", ascending=False)
        .head(10)
        .rename(columns={"comuna": "Comuna"})
    )
    df_com["% del total"] = (df_com["Total"] / total * 100).round(1).astype(str) + "%"
    story.append(_df_to_table(df_com, col_widths=[8*cm, 3*cm, 4*cm]))
    story.append(Spacer(1, 0.3*cm))

    # Nuevos desistes
    if not nuevos.empty:
        story.append(PageBreak())
        _sigma_header(story, "Nuevos Desistes Detectados", "", corte)
        story.append(Paragraph(
            f"Se registraron {len(nuevos)} nuevos desistes desde el corte anterior.",
            s["normal"],
        ))
        story.append(Spacer(1, 0.2*cm))
        cols = [c for c in ["rut_norm","nombre","course_code","specialty","sexo","comuna","edad","is_repeat"]
                if c in nuevos.columns]
        story.append(_df_to_table(
            nuevos[cols].reset_index(drop=True),
            col_widths=[2.8*cm, 5*cm, 1.8*cm, 2.5*cm, 2.2*cm, 2.5*cm, 1.5*cm, 1.7*cm],
        ))

    # Nómina completa
    story.append(PageBreak())
    _sigma_header(story, "Nómina Completa de Desistes", "", corte)
    cols_nom = [c for c in ["rut_norm","nombre","course_code","specialty","sexo","comuna","edad","is_repeat"]
                if c in dd.columns]
    dd_sorted = dd.sort_values(["specialty", "level", "course_code"])
    story.append(_df_to_table(
        dd_sorted[cols_nom].reset_index(drop=True),
        col_widths=[2.8*cm, 5*cm, 1.8*cm, 2.5*cm, 2.2*cm, 2.5*cm, 1.5*cm, 1.7*cm],
    ))

    # ── Gráficos ──────────────────────────────────────────────────────
    if _CHARTS_OK:
        story.append(PageBreak())
        _sigma_header(story, "Análisis Visual de Desistimiento", "", corte)

        story.append(_section_title("Tasa de desistimiento por nivel"))
        img, txt = pdf_chart_des_nivel(dd, df_current)
        _chart_block(story, img, txt)

        story.append(_section_title("Tasa de desistimiento por especialidad"))
        img, txt = pdf_chart_des_especialidad(dd, df_current)
        _chart_block(story, img, txt)

        # Perfil: repitentes + género
        _sx_dd = _norm_sexo(dd["sexo"]) if "sexo" in dd.columns else pd.Series(dtype=str)
        n_h_des = int((_sx_dd == "M").sum())
        n_m_des = int((_sx_dd == "F").sum())
        n_rep_des = int(dd["is_repeat"].sum()) if "is_repeat" in dd.columns else 0
        n_norep_des = total - n_rep_des
        story.append(_section_title("Perfil de quienes abandonan"))
        img, txt = pdf_chart_des_perfil(n_rep_des, n_norep_des, n_h_des, n_m_des)
        _chart_block(story, img, txt)

    _sigma_footer(story)
    doc.build(story)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════
# 3. REPORTE POR ESPECIALIDAD — PDF
# ═════════════════════════════════════════════════════════════════════
def generate_pdf_especialidad(
    df_current: pd.DataFrame,
    df_desiste: Optional[pd.DataFrame],
    corte: str,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    s = _build_styles()
    story = []

    df = df_current.copy()
    df["specialty"]    = df.get("specialty", pd.Series(dtype=str)).fillna("SIN DATOS").str.upper()
    # Normalizar sexo a M/F independiente del formato fuente
    df["_sx"]          = _norm_sexo(df["sexo"]) if "sexo" in df.columns else ""
    df["is_repeat"]    = df.get("is_repeat", pd.Series(dtype=bool)).fillna(False).astype(bool)
    df["is_extranjero"]= ~df.get("nacionalidad", pd.Series(dtype=str)).fillna("CHILENA").str.upper().str.strip().isin(
                            ["CHILENA","CHILENO","CHILE","NAN",""])
    df["level"]        = pd.to_numeric(df.get("level", pd.Series(dtype=int)), errors="coerce").fillna(0).astype(int)

    specs = sorted([s_ for s_ in df["specialty"].unique() if s_ not in ("SIN DATOS","")])
    total = len(df)

    # Pre-calcular desistes por especialidad
    des_por_spec = {}
    if df_desiste is not None and not df_desiste.empty:
        dd_ = df_desiste.copy()
        dd_["specialty"] = dd_.get("specialty", pd.Series(dtype=str)).fillna("").str.upper()
        for sp in specs:
            des_por_spec[sp] = int((dd_["specialty"] == sp).sum())

    _sigma_header(story, "Reporte por Especialidad", "", corte)

    # ── 1. Resumen consolidado ─────────────────────────────────────────
    story.append(_section_title("Resumen por especialidad"))
    df_res = pd.DataFrame([{
        "Especialidad":  sp,
        "Total":         len(df[df["specialty"]==sp]),
        "Hombres":       int((df[df["specialty"]==sp]["_sx"]=="M").sum()),
        "Mujeres":       int((df[df["specialty"]==sp]["_sx"]=="F").sum()),
        "Extranjeros":   int(df[df["specialty"]==sp]["is_extranjero"].sum()),
        "Repitentes":    int(df[df["specialty"]==sp]["is_repeat"].sum()),
        "Desistes":      des_por_spec.get(sp, 0),
        "% matrícula":   str(round(len(df[df["specialty"]==sp])/total*100,1))+"%",
    } for sp in specs])
    story.append(_df_to_table(
        df_res,
        col_widths=[3.8*cm, 1.8*cm, 2.2*cm, 2.2*cm, 2.5*cm, 2.5*cm, 2.2*cm, 2.3*cm],
    ))
    story.append(Spacer(1, 0.5*cm))

    # ── 2. Detalle por curso ───────────────────────────────────────────
    story.append(_section_title("Detalle por curso"))
    df_cur = (
        df.groupby(["course_code", "specialty", "level"], dropna=False)
        .agg(
            Total       = ("rut_norm",      "count"),
            Hombres     = ("_sx",           lambda x: (x=="M").sum()),
            Mujeres     = ("_sx",           lambda x: (x=="F").sum()),
            Extranjeros = ("is_extranjero", "sum"),
            Repitentes  = ("is_repeat",     "sum"),
        )
        .reset_index()
        .sort_values(["specialty", "level", "course_code"])
        .rename(columns={"course_code": "Curso", "specialty": "Especialidad", "level": "Nivel"})
    )
    story.append(_df_to_table(
        df_cur,
        col_widths=[2.5*cm, 3.8*cm, 1.5*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm],
    ))
    story.append(Spacer(1, 0.5*cm))

    # ── 3. Indicadores por especialidad (una fila de KPIs por esp.) ────
    story.append(_section_title("Indicadores por especialidad"))
    story.append(Spacer(1, 0.3*cm))
    for sp in specs:
        color = SPEC_COLORS.get(sp, C_MUTED)
        dfs   = df[df["specialty"]==sp]
        n     = len(dfs)
        n_h   = int((dfs["_sx"]=="M").sum())
        n_m   = int((dfs["_sx"]=="F").sum())
        n_ext = int(dfs["is_extranjero"].sum())
        n_rep = int(dfs["is_repeat"].sum())
        n_des = des_por_spec.get(sp, 0)
        pct_rep = f"{round(n_rep/n*100,1)}%" if n > 0 else "0%"
        pct_des = f"{round(n_des/n*100,1)}%" if n > 0 else "0%"

        # Etiqueta de especialidad como mini-título
        story.append(Table(
            [[Paragraph(sp, ParagraphStyle(
                f"sp_{sp}", fontName="Helvetica-Bold", fontSize=9,
                textColor=color, leading=11,
            ))]],
            colWidths=[A4[0] - 3*cm], rowHeights=[16],
        ))
        story.append(_kpi_table([
            ("Matriculados",  n,     color),
            ("Hombres",       n_h,   C_CYAN),
            ("Mujeres",       n_m,   C_GREEN),
            ("Extranjeros",   n_ext, C_AMBER),
            ("Repitentes",    n_rep, C_AMBER, pct_rep),
            ("Desistes",      n_des, C_RED,   pct_des),
        ]))
        story.append(Spacer(1, 0.4*cm))

    # ── Gráficos ──────────────────────────────────────────────────────
    if _CHARTS_OK:
        story.append(PageBreak())
        _sigma_header(story, "Análisis Visual por Especialidad", "", corte)

        story.append(_section_title("Distribución por curso y especialidad"))
        img, txt = pdf_chart_cursos(df)
        _chart_block(story, img, txt)

        story.append(_section_title("Indicadores de contexto por especialidad"))
        img, txt = pdf_chart_heatmap(df, df_desiste)
        _chart_block(story, img, txt)

    _sigma_footer(story)
    doc.build(story)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════
# 4. EXCEL EJECUTIVO
# ═════════════════════════════════════════════════════════════════════
def _xl_header(ws, titulo: str, corte: str, start_row: int = 1) -> int:
    """Escribe header SIGMA en la hoja. Retorna la fila siguiente."""
    ws.merge_cells(f"A{start_row}:H{start_row}")
    c = ws.cell(start_row, 1, "SIGMA")
    c.font = Font(name="Calibri", bold=True, size=20, color="1E3A5F")
    c.alignment = Alignment(horizontal="left")

    ws.merge_cells(f"A{start_row+1}:H{start_row+1}")
    c2 = ws.cell(start_row+1, 1, titulo)
    c2.font = Font(name="Calibri", bold=True, size=14, color="111827")

    ws.merge_cells(f"A{start_row+2}:H{start_row+2}")
    c3 = ws.cell(start_row+2, 1, f"Corte: {corte}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c3.font = Font(name="Calibri", size=9, color="6B7280")

    for row in range(start_row, start_row+3):
        ws.row_dimensions[row].height = 22 if row == start_row else 18
        for col in range(1, 9):
            ws.cell(row, col).fill = PatternFill("solid", fgColor="FFFFFF")

    return start_row + 4


def _xl_write_df(ws, df: pd.DataFrame, start_row: int,
                 header_color: str = "63B3ED") -> int:
    thin = Side(style="thin", color="4A5568")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Encabezados
    for col_idx, col_name in enumerate(df.columns, 1):
        c = ws.cell(start_row, col_idx, str(col_name).upper())
        c.font      = Font(name="Calibri", bold=True, size=9, color="080C14")
        c.fill      = PatternFill("solid", fgColor=header_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
        ws.row_dimensions[start_row].height = 20

    # Datos
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        bg = "F9FAFB" if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row_idx, col_idx, val if pd.notna(val) else "")
            c.font      = Font(name="Calibri", size=9, color="111827")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border    = border
        ws.row_dimensions[row_idx].height = 16

    # Autofit columnas
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx-1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    return start_row + len(df) + 2


def generate_excel_ejecutivo(
    metrics: dict,
    df_current: pd.DataFrame,
    df_desiste: Optional[pd.DataFrame],
    df_diff: Optional[pd.DataFrame],
    corte: str,
) -> bytes:
    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "63B3ED"

    row = _xl_header(ws1, "Resumen Ejecutivo de Matrícula", corte)

    kpis = [
        ("Matriculados",    metrics.get("matriculados", 0),    "63B3ED"),
        ("Desiste",         metrics.get("desiste", 0),         "FC8181"),
        ("Total esperado",  metrics.get("total_esperado", 0),  "F6AD55"),
        ("Hombres",         metrics.get("sexo_h", 0),          "63B3ED"),
        ("Mujeres",         metrics.get("sexo_m", 0),          "9AE6B4"),
        ("Repitentes",      metrics.get("repitentes", 0),      "F6AD55"),
        ("% Renca",         str(metrics.get("pct_renca", 0))+"%", "68D391"),
        ("Edad promedio",   metrics.get("edad_prom", 0),       "76E4F7"),
    ]
    for i, (label, value, color) in enumerate(kpis):
        col = (i % 4) + 1
        if i % 4 == 0 and i > 0:
            row += 3
        ws1.cell(row,     col, str(value)).font = Font(name="Calibri", bold=True, size=18, color=color)
        ws1.cell(row+1,   col, label).font      = Font(name="Calibri", size=8, color="6B7280")
        ws1.cell(row,     col).fill = PatternFill("solid", fgColor="FFFFFF")
        ws1.cell(row+1,   col).fill = PatternFill("solid", fgColor="FFFFFF")
        ws1.column_dimensions[get_column_letter(col)].width = 20
        ws1.row_dimensions[row].height   = 28
        ws1.row_dimensions[row+1].height = 14

    row += 5

    # Distribución por especialidad
    ws1.cell(row, 1, "DISTRIBUCIÓN POR ESPECIALIDAD").font = Font(
        name="Calibri", bold=True, size=11, color="63B3ED")
    ws1.cell(row, 1).fill = PatternFill("solid", fgColor="FFFFFF")
    row += 1

    if df_current is not None and not df_current.empty:
        df_sp = (
            df_current.groupby("specialty")
            .agg(
                Total      = ("rut_norm", "count"),
                Hombres    = ("sexo", lambda x: (x.str.upper()=="M").sum()),
                Mujeres    = ("sexo", lambda x: (x.str.upper()=="F").sum()),
            )
            .reset_index()
            .sort_values("Total", ascending=False)
            .rename(columns={"specialty": "Especialidad"})
        )
        df_sp["% Total"] = (df_sp["Total"] / len(df_current) * 100).round(1).astype(str) + "%"
        row = _xl_write_df(ws1, df_sp, row)

    # ── Hoja 2: Desistimiento ────────────────────────────────────────
    ws2 = wb.create_sheet("Desistimiento")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "FC8181"

    row2 = _xl_header(ws2, "Análisis de Desistimiento", corte)

    if df_desiste is not None and not df_desiste.empty:
        dd = df_desiste.copy()
        dd["specialty"] = dd.get("specialty", pd.Series(dtype=str)).fillna("").str.upper()
        dd["is_repeat"] = dd.get("is_repeat", pd.Series(dtype=bool)).fillna(False).astype(bool)

        # Resumen por especialidad
        ws2.cell(row2, 1, "POR ESPECIALIDAD").font = Font(
            name="Calibri", bold=True, size=11, color="FC8181")
        ws2.cell(row2, 1).fill = PatternFill("solid", fgColor="FFFFFF")
        row2 += 1

        df_des_sp = (
            dd.groupby("specialty")
            .agg(
                Total      = ("rut_norm", "count"),
                Repitentes = ("is_repeat", "sum"),
                Hombres    = ("sexo",      lambda x: (x.str.upper()=="MASCULINO").sum()),
                Mujeres    = ("sexo",      lambda x: (x.str.upper()=="FEMENINO").sum()),
            )
            .reset_index()
            .sort_values("Total", ascending=False)
            .rename(columns={"specialty": "Especialidad"})
        )
        df_des_sp["% Repitentes"] = (df_des_sp["Repitentes"] / df_des_sp["Total"] * 100).round(1).astype(str) + "%"
        row2 = _xl_write_df(ws2, df_des_sp, row2, header_color="FC8181")

        # Nómina completa
        ws2.cell(row2, 1, "NÓMINA COMPLETA").font = Font(
            name="Calibri", bold=True, size=11, color="FC8181")
        ws2.cell(row2, 1).fill = PatternFill("solid", fgColor="FFFFFF")
        row2 += 1

        cols_nom = [c for c in ["rut_norm","nombre","course_code","specialty",
                                "sexo","comuna","edad","is_repeat","nacionalidad"]
                    if c in dd.columns]
        row2 = _xl_write_df(ws2, dd[cols_nom].sort_values("specialty"), row2, header_color="FC8181")

    # ── Hoja 3: Especialidades ───────────────────────────────────────
    ws3 = wb.create_sheet("Especialidades")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "F6AD55"

    row3 = _xl_header(ws3, "Reporte por Especialidad", corte)

    if df_current is not None and not df_current.empty:
        df3 = df_current.copy()
        df3["specialty"]    = df3.get("specialty", pd.Series(dtype=str)).fillna("").str.upper()
        df3["is_repeat"]    = df3.get("is_repeat", pd.Series(dtype=bool)).fillna(False).astype(bool)
        df3["is_extranjero"]= ~df3.get("nacionalidad", pd.Series(dtype=str)).fillna("").str.upper().isin(
                                ["CHILENA","CHILENO","CHILE",""])
        df3["level"]        = pd.to_numeric(df3.get("level", pd.Series(dtype=int)),
                                errors="coerce").fillna(0).astype(int)

        df_res3 = pd.DataFrame([{
            "Especialidad": sp,
            "Total":        len(df3[df3["specialty"]==sp]),
            "Hombres":      int((df3[df3["specialty"]==sp]["sexo"].str.upper()=="M").sum()),
            "Mujeres":      int((df3[df3["specialty"]==sp]["sexo"].str.upper()=="F").sum()),
            "Extranjeros":  int(df3[df3["specialty"]==sp]["is_extranjero"].sum()),
            "Repitentes":   int(df3[df3["specialty"]==sp]["is_repeat"].sum()),
            "% Total":      str(round(len(df3[df3["specialty"]==sp])/len(df3)*100,1))+"%",
        } for sp in sorted(df3["specialty"].unique()) if sp])

        ws3.cell(row3, 1, "RESUMEN POR ESPECIALIDAD").font = Font(
            name="Calibri", bold=True, size=11, color="F6AD55")
        ws3.cell(row3, 1).fill = PatternFill("solid", fgColor="FFFFFF")
        row3 += 1
        row3 = _xl_write_df(ws3, df_res3, row3, header_color="F6AD55")

        # Detalle por curso
        df_cur3 = (
            df3.groupby(["course_code","specialty","level"], dropna=False)
            .agg(
                Total       = ("rut_norm",      "count"),
                Hombres     = ("sexo",          lambda x: (x.str.upper()=="M").sum()),
                Mujeres     = ("sexo",          lambda x: (x.str.upper()=="F").sum()),
                Extranjeros = ("is_extranjero", "sum"),
                Repitentes  = ("is_repeat",     "sum"),
            )
            .reset_index()
            .sort_values(["level","course_code"])
            .rename(columns={"course_code":"Curso","specialty":"Especialidad","level":"Nivel"})
        )
        ws3.cell(row3, 1, "DETALLE POR CURSO").font = Font(
            name="Calibri", bold=True, size=11, color="F6AD55")
        ws3.cell(row3, 1).fill = PatternFill("solid", fgColor="FFFFFF")
        row3 += 1
        row3 = _xl_write_df(ws3, df_cur3, row3, header_color="F6AD55")

    # ── Fondo oscuro general ─────────────────────────────────────────
    for ws in [ws1, ws2, ws3]:
        ws.sheet_view.showGridLines = False
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.fill.fgColor.rgb in ("00000000", ""):
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()